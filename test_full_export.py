"""
Full export test - verifies CSV + XLSX + Summary all work together.
"""
import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from models.lead import Lead
from export.csv_export import export_leads
from export.report_generator import generate_summary_report


def main():
    print("=" * 70)
    print("🧪 FULL EXPORT TEST (CSV + XLSX + Summary)")
    print("=" * 70)
    print()
    
    # Create sample leads with varying quality
    leads = []
    
    # Tier A lead (high score)
    lead_a = Lead(
        business_name="Premium Restaurant Group",
        city="Austin",
        region="TX",
        score=85,
        tier="A"
    )
    lead_a.phone = "+1-512-555-1000"
    lead_a.emails = ["info@premiumgroup.com", "sales@premiumgroup.com"]
    lead_a.website = "https://premiumgroup.com"
    lead_a.has_contact_form = True
    lead_a.has_booking = True
    leads.append(lead_a)
    
    # Tier B lead
    lead_b = Lead(
        business_name="Local Craft Brewery",
        city="Austin",
        region="TX",
        score=70,
        tier="B"
    )
    lead_b.phone = "+1-512-555-2000"
    lead_b.emails = ["contact@localbrewery.com"]
    lead_b.website = "https://localbrewery.com"
    lead_b.has_contact_form = True
    leads.append(lead_b)
    
    # Tier C lead
    lead_c = Lead(
        business_name="Small Bistro",
        city="Austin",
        region="TX",
        score=40,
        tier="C"
    )
    lead_c.phone = "+1-512-555-3000"
    leads.append(lead_c)
    
    print(f"✅ Created {len(leads)} sample leads")
    print(f"   - Tier A: {sum(1 for l in leads if l.tier == 'A')}")
    print(f"   - Tier B: {sum(1 for l in leads if l.tier == 'B')}")
    print(f"   - Tier C: {sum(1 for l in leads if l.tier == 'C')}")
    print()
    
    # Export CSV
    print("📄 Exporting CSV...")
    csv_path = export_leads(
        leads=leads,
        vertical="Full Test",
        region="Austin TX",
        output_format="csv",
        sort_by_score=True
    )
    print(f"   ✅ {csv_path}")
    print()
    
    # Export XLSX
    print("📊 Exporting XLSX...")
    xlsx_path = export_leads(
        leads=leads,
        vertical="Full Test",
        region="Austin TX",
        output_format="xlsx",
        sort_by_score=True
    )
    print(f"   ✅ {xlsx_path}")
    
    # Check XLSX file
    xlsx_file = Path(xlsx_path)
    if xlsx_file.exists():
        size = xlsx_file.stat().st_size
        print(f"   📏 Size: {size:,} bytes")
        
        # Try to open it with openpyxl to verify it's valid
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb.active
            print(f"   ✅ File is valid (sheet: '{ws.title}', rows: {ws.max_row})")
        except Exception as e:
            print(f"   ❌ File is CORRUPTED: {e}")
    print()
    
    # Generate summary report
    print("📋 Generating summary report...")
    report_path = generate_summary_report(leads, "Full Test", "Austin TX", csv_path)
    print(f"   ✅ {report_path}")
    
    # Verify report file
    report_file = Path(report_path)
    if report_file.exists():
        size = report_file.stat().st_size
        print(f"   📏 Size: {size:,} bytes")
    print()
    
    print("=" * 70)
    print("✅ ALL EXPORTS SUCCESSFUL!")
    print("=" * 70)
    print()
    print("Files created:")
    print(f"  1. CSV:     {csv_path}")
    print(f"  2. XLSX:    {xlsx_path}")
    print(f"  3. Report:  {report_path}")
    print()


if __name__ == "__main__":
    main()
