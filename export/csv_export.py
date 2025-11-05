"""
CSV export functionality for lead data.
Professional B2B lead generation export with clean formatting and standardized schema.
"""
from typing import List, Optional, Dict, Literal, Tuple
from datetime import datetime
from collections import OrderedDict
import pandas as pd
from pathlib import Path
import re
from models.lead import Lead
from config.settings import SETTINGS
from utils.logging_utils import get_logger

log = get_logger(__name__)

# Export format type
ExportFormat = Literal["csv", "xlsx"]


# ============================================================================
# STANDARDIZED B2B LEAD SCHEMA
# ============================================================================

def get_lead_schema() -> OrderedDict:
    """
    Define the standardized B2B lead generation CSV schema.
    
    Returns:
        OrderedDict mapping internal field names (snake_case) to display names (Title Case).
        The order defines the column order in the exported CSV/XLSX.
    """
    return OrderedDict([
        # ========== CONTACT INFORMATION ==========
        ('company_name', 'Company Name'),
        ('website', 'Website'),
        ('contact_name', 'Contact Name'),
        ('job_title', 'Job Title'),
        ('email', 'Email'),
        ('phone_number', 'Phone Number'),
        ('linkedin_url', 'LinkedIn URL'),
        
        # ========== COMPANY DETAILS ==========
        ('industry', 'Industry'),
        ('employee_count', 'Employee Count'),
        ('estimated_revenue', 'Estimated Revenue'),
        ('headquarters_location', 'Headquarters Location'),
        ('tech_stack', 'Tech Stack'),
        
        # ========== LEAD INTELLIGENCE ==========
        ('pain_point_indicator', 'Pain Point / Indicator'),
        ('recent_activity', 'Recent Activity'),
        ('lead_source', 'Lead Source'),
        ('confidence_score', 'Confidence Score'),
        ('date_added', 'Date Added'),
        ('notes_hook', 'Notes / Personalized Hook'),
        
        # ========== SUPPLEMENTARY DATA (Optional) ==========
        ('yelp_rating', 'Yelp Rating'),
        ('yelp_reviews', 'Yelp Reviews'),
        ('google_rating', 'Google Rating'),
        ('google_reviews', 'Google Reviews'),
        ('has_contact_form', 'Has Contact Form'),
        ('has_online_booking', 'Has Online Booking'),
        ('uses_https', 'Uses HTTPS'),
        ('tier', 'Tier'),
        ('internal_id', 'Internal ID'),
    ])


# ============================================================================
# DATA NORMALIZATION & FORMATTING
# ============================================================================

def _normalize_phone(phone: Optional[str]) -> str:
    """
    Normalize phone number to clean, consistent format.
    
    Returns:
        Formatted phone string or empty string if invalid.
    """
    if not phone:
        return ""
    
    # Strip all non-digit characters
    digits = re.sub(r'\D', '', phone)
    
    # Format as E.164 for North America (+1-XXX-XXX-XXXX)
    if len(digits) == 10:
        return f"+1-{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    elif len(digits) == 11 and digits[0] == '1':
        return f"+1-{digits[1:4]}-{digits[4:7]}-{digits[7:]}"
    else:
        # Return as-is if format is unknown (international, etc.)
        return phone


def _normalize_website(website: Optional[str]) -> str:
    """Ensure website has proper URL scheme."""
    if not website:
        return ""
    
    website = website.strip()
    if not website.startswith(('http://', 'https://')):
        return f"https://{website}"
    return website


def _format_boolean(value: Optional[bool]) -> str:
    """Format boolean for CSV display."""
    if value is None:
        return ""
    return "Yes" if value else "No"


def _format_date(date_str: str) -> str:
    """Format ISO date to human-readable format (YYYY-MM-DD HH:MM UTC)."""
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime("%Y-%m-%d %H:%M UTC")
    except:
        return datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")


def _extract_pain_points(lead: Lead) -> str:
    """
    Analyze lead data to extract pain points or opportunity indicators.
    
    Returns:
        Human-readable pain point description or empty string.
    """
    pain_points = []
    
    if lead.has_contact_form is False:
        pain_points.append("Missing contact form")
    
    if lead.has_booking is False:
        pain_points.append("No online booking")
    
    if lead.uses_https is False:
        pain_points.append("No HTTPS (security risk)")
    
    if lead.has_financing is False and lead.yelp_price_level in ['$$$', '$$$$']:
        pain_points.append("High-value service without financing options")
    
    if lead.tech_stack and any(tech.lower() in ['wordpress', 'wix', 'squarespace'] for tech in lead.tech_stack):
        pain_points.append("Legacy website platform")
    
    if not pain_points and lead.score >= 80:
        return "High-quality lead - general outreach opportunity"
    
    return " | ".join(pain_points) if pain_points else ""


def _extract_recent_activity(lead: Lead) -> str:
    """Extract recent activity indicators from lead data."""
    if lead.tavily_recent_activity:
        return "Recent online mentions found"
    
    if lead.tavily_sources_found > 3:
        return f"{lead.tavily_sources_found} online sources found"
    
    if lead.google_review_count and lead.google_review_count > 50:
        return f"{lead.google_review_count} Google reviews (active presence)"
    
    return ""


def _map_lead_to_schema(lead: Lead) -> Dict[str, any]:
    """
    Map a Lead object to the standardized B2B schema.
    
    This is the core transformation function that converts internal lead data
    to the clean, client-facing B2B format defined in get_lead_schema().
    
    Args:
        lead: Lead object with all enrichment data
    
    Returns:
        Dictionary with keys matching the schema's internal field names
    """
    # Get primary email
    primary_email = lead.emails[0] if lead.emails else ""
    
    # Extract location (city, region)
    location_parts = [lead.city, lead.region]
    headquarters = ", ".join(filter(None, location_parts)) or ""
    
    # Format tech stack
    tech_stack_str = ", ".join(lead.tech_stack) if lead.tech_stack else ""
    
    # Build notes/hook from multiple sources
    notes_parts = []
    
    # Add pain point context
    pain_point = _extract_pain_points(lead)
    if pain_point:
        notes_parts.append(f"Opportunity: {pain_point}")
    
    # Add any internal notes
    if lead.internal_notes:
        notes_parts.append(lead.internal_notes)
    
    # Add outreach snippet if available
    if lead.outreach_snippet:
        notes_parts.append(f"Hook: {lead.outreach_snippet}")
    
    personalized_hook = " | ".join(notes_parts) if notes_parts else ""
    
    # Determine lead source
    lead_source = lead.discovery_method or lead.source or "Unknown"
    
    # Determine industry from categories
    industry = ""
    if lead.yelp_categories:
        industry = lead.yelp_categories[0]  # Use primary category
    
    return {
        # ========== CONTACT INFORMATION ==========
        'company_name': lead.business_name,
        'website': _normalize_website(lead.website),
        'contact_name': "",  # Not yet populated - reserved for future enrichment
        'job_title': "",  # Not yet populated - reserved for future enrichment
        'email': primary_email,
        'phone_number': _normalize_phone(lead.phone),
        'linkedin_url': "",  # Not yet populated - reserved for future enrichment
        
        # ========== COMPANY DETAILS ==========
        'industry': industry,
        'employee_count': "",  # Not yet populated - reserved for future enrichment
        'estimated_revenue': "",  # Not yet populated - reserved for future enrichment
        'headquarters_location': headquarters,
        'tech_stack': tech_stack_str,
        
        # ========== LEAD INTELLIGENCE ==========
        'pain_point_indicator': pain_point,
        'recent_activity': _extract_recent_activity(lead),
        'lead_source': lead_source,
        'confidence_score': int(lead.score),  # Convert to int for cleaner display
        'date_added': _format_date(lead.scrape_date),
        'notes_hook': personalized_hook,
        
        # ========== SUPPLEMENTARY DATA ==========
        'yelp_rating': lead.yelp_rating if lead.yelp_rating and lead.yelp_rating > 0 else "",
        'yelp_reviews': lead.yelp_review_count if lead.yelp_review_count else "",
        'google_rating': lead.google_rating if lead.google_rating and lead.google_rating > 0 else "",
        'google_reviews': lead.google_review_count if lead.google_review_count else "",
        'has_contact_form': _format_boolean(lead.has_contact_form),
        'has_online_booking': _format_boolean(lead.has_booking),
        'uses_https': _format_boolean(lead.uses_https),
        'tier': lead.tier or "",
        'internal_id': lead.lead_id,
    }


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_leads(
    leads: List[Lead],
    vertical: str,
    region: str,
    output_format: ExportFormat = "csv",
    sort_by_score: bool = True
) -> str:
    """
    Export leads to CSV or XLSX with professional B2B formatting.
    
    Args:
        leads: List of Lead objects to export
        vertical: Business vertical (e.g., "HVAC", "SaaS Companies")
        region: Geographic region (e.g., "Milton, Ontario")
        output_format: Export format - "csv" or "xlsx" (default: "csv")
        sort_by_score: Sort leads by confidence score descending (default: True)
    
    Returns:
        str: Absolute path to the created file
    
    Example:
        >>> leads = [lead1, lead2, lead3]
        >>> path = export_leads(leads, "Digital Agencies", "Toronto, ON")
        >>> print(f"Exported to: {path}")
    """
    # Validate inputs
    if not leads:
        log.warning("No leads to export")
        return ""
    
    # Sort leads by confidence score if requested
    if sort_by_score:
        leads = sorted(leads, key=lambda x: x.score, reverse=True)
    
    # Get schema definition
    schema = get_lead_schema()
    
    # Transform leads to standardized format
    log.info(f"Transforming {len(leads)} leads to B2B schema...")
    formatted_data = [_map_lead_to_schema(lead) for lead in leads]
    
    # Create DataFrame with exact column order from schema
    df = pd.DataFrame(formatted_data)
    
    # Ensure all schema columns exist (add empty columns if missing)
    for field_name in schema.keys():
        if field_name not in df.columns:
            df[field_name] = ""
    
    # Reorder columns to match schema
    df = df[list(schema.keys())]
    
    # Rename columns to display names (Title Case headers)
    df.columns = [schema[col] for col in df.columns]
    
    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Clean vertical and region for filename
    vertical_clean = re.sub(r'[^\w\s-]', '', vertical).replace(" ", "_")
    region_clean = re.sub(r'[^\w\s-]', '', region).replace(" ", "_")
    
    # Create filename with appropriate extension
    file_extension = output_format.lower()
    filename = f"{vertical_clean}_{region_clean}_{timestamp}.{file_extension}"
    filepath = Path(SETTINGS.OUTPUT_DIR) / filename
    
    # Ensure output directory exists
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    # Export based on format
    if output_format == "xlsx":
        _export_to_xlsx(df, filepath)
    else:
        _export_to_csv(df, filepath)
    
    log.info(f"✅ Exported {len(leads)} leads to: {filepath}")
    log.info(f"   Format: {output_format.upper()} | Sorted by score: {sort_by_score}")
    
    return str(filepath)


def _export_to_csv(df: pd.DataFrame, filepath: Path) -> None:
    """Export DataFrame to CSV with UTF-8 encoding."""
    df.to_csv(
        filepath,
        index=False,
        encoding='utf-8',
        lineterminator='\n'  # Use Unix line endings for consistency
    )


def _export_to_xlsx(df: pd.DataFrame, filepath: Path) -> None:
    """
    Export DataFrame to XLSX with professional formatting.
    Includes auto-column width, frozen header row, and filter dropdowns.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed - falling back to basic XLSX export")
        df.to_excel(filepath, index=False, engine='openpyxl')
        return
    
    # Create Excel writer
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Leads', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Leads']
        
        # Style the header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths based on content
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with padding (max 50 chars)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Add auto-filter to all columns
        worksheet.auto_filter.ref = worksheet.dimensions
    
    log.info("   Applied XLSX styling: header formatting, auto-width, filters")


# ============================================================================
# BACKWARD COMPATIBILITY
# ============================================================================

def export_to_csv(leads: List[Lead], vertical: str, region: str) -> str:
    """
    Backward-compatible wrapper for export_leads().
    Exports to CSV format only.
    
    Args:
        leads: List of Lead objects
        vertical: Business vertical
        region: Geographic region
    
    Returns:
        Path to created CSV file
    """
    return export_leads(leads, vertical, region, output_format="csv")


# ============================================================================
# STATISTICS & REPORTING
# ============================================================================

def get_export_stats(leads: List[Lead]) -> dict:
    """
    Calculate comprehensive export statistics.
    
    Args:
        leads: List of Lead objects
    
    Returns:
        Dictionary with detailed statistics about the lead set
    """
    if not leads:
        return {
            'total_leads': 0,
            'tier_a': 0,
            'tier_b': 0,
            'tier_c': 0,
            'avg_score': 0.0,
            'with_email': 0,
            'with_phone': 0,
            'with_website': 0,
            'with_form': 0,
            'with_booking': 0,
            'with_https': 0,
        }
    
    tier_counts = {"A": 0, "B": 0, "C": 0}
    
    for lead in leads:
        if lead.tier:
            tier_counts[lead.tier] += 1
    
    # Calculate metrics
    total_leads = len(leads)
    avg_score = sum(lead.score for lead in leads) / total_leads
    
    # Count enrichment signals
    with_email = sum(1 for lead in leads if lead.emails)
    with_phone = sum(1 for lead in leads if lead.phone)
    with_website = sum(1 for lead in leads if lead.website)
    with_form = sum(1 for lead in leads if lead.has_contact_form is True)
    with_booking = sum(1 for lead in leads if lead.has_booking is True)
    with_https = sum(1 for lead in leads if lead.uses_https is True)
    
    return {
        'total_leads': total_leads,
        'tier_a': tier_counts['A'],
        'tier_b': tier_counts['B'],
        'tier_c': tier_counts['C'],
        'avg_score': round(avg_score, 1),
        'with_email': with_email,
        'with_phone': with_phone,
        'with_website': with_website,
        'with_form': with_form,
        'with_booking': with_booking,
        'with_https': with_https,
    }


def print_export_summary(leads: List[Lead], filepath: str) -> None:
    """
    Print a formatted summary of the export.
    
    Args:
        leads: List of exported leads
        filepath: Path to the exported file
    """
    stats = get_export_stats(leads)
    
    print("\n" + "="*60)
    print("📊 EXPORT SUMMARY")
    print("="*60)
    print(f"📄 File: {Path(filepath).name}")
    print(f"📍 Location: {filepath}")
    print(f"\n🎯 LEAD METRICS:")
    print(f"   Total Leads: {stats['total_leads']}")
    print(f"   Average Score: {stats['avg_score']}")
    print(f"\n🏆 TIER BREAKDOWN:")
    print(f"   Tier A: {stats['tier_a']}")
    print(f"   Tier B: {stats['tier_b']}")
    print(f"   Tier C: {stats['tier_c']}")
    print(f"\n✅ ENRICHMENT COVERAGE:")
    print(f"   With Email: {stats['with_email']} ({stats['with_email']/stats['total_leads']*100:.1f}%)")
    print(f"   With Phone: {stats['with_phone']} ({stats['with_phone']/stats['total_leads']*100:.1f}%)")
    print(f"   With Website: {stats['with_website']} ({stats['with_website']/stats['total_leads']*100:.1f}%)")
    print(f"   With Contact Form: {stats['with_form']} ({stats['with_form']/stats['total_leads']*100:.1f}%)")
    print(f"   With Booking: {stats['with_booking']} ({stats['with_booking']/stats['total_leads']*100:.1f}%)")
    print(f"   Uses HTTPS: {stats['with_https']} ({stats['with_https']/stats['total_leads']*100:.1f}%)")
    print("="*60 + "\n")
